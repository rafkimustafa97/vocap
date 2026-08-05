import openpyxl
import json
import os

excel_path = r'c:\Users\rafki\Documents\vocap\source\Skema_Belajar_5Bulan_Lengkap.xlsx'
target_ts_path = r'c:\Users\rafki\Documents\vocap\src\data\wordsMaster.ts'

print("Loading Excel workbook...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

master_ws = wb['Master Kata (Bernomor)']
afiks_ws = wb['Jadwal Mingguan (Afiks)']
wf_ws = wb['Word Family & Kolokasi']

# 1. Map Afiks sheet
afiks_map = {}
for row in list(afiks_ws.iter_rows(values_only=True))[1:]:
    if not row or len(row) < 4 or not row[2]:
        continue
    base = str(row[2]).strip().lower()
    derived = str(row[3]).strip() if row[3] else ''
    affix_type = str(row[4]).strip() if len(row) > 4 and row[4] else ''
    affix_used = str(row[5]).strip() if len(row) > 5 and row[5] else ''
    pos = str(row[6]).strip() if len(row) > 6 and row[6] else ''
    
    if base not in afiks_map:
        afiks_map[base] = []
    afiks_map[base].append({'derived': derived, 'type': affix_type, 'used': affix_used, 'pos': pos})

# 2. Map Word Family sheet
wf_map = {}
for row in list(wf_ws.iter_rows(values_only=True))[1:]:
    if not row or len(row) < 2 or not row[1]:
        continue
    word = str(row[1]).strip().lower()
    pos = str(row[2]).strip() if len(row) > 2 and row[2] else ''
    noun_f = str(row[3]).strip() if len(row) > 3 and row[3] else '-'
    verb_f = str(row[4]).strip() if len(row) > 4 and row[4] else '-'
    adj_f = str(row[5]).strip() if len(row) > 5 and row[5] else '-'
    adv_f = str(row[6]).strip() if len(row) > 6 and row[6] else '-'
    colloc = str(row[7]).strip() if len(row) > 7 and row[7] else '-'
    
    wf_map[word] = {
        'pos': pos,
        'noun_f': noun_f,
        'verb_f': verb_f,
        'adj_f': adj_f,
        'adv_f': adv_f,
        'colloc': colloc
    }

# 3. Process Master Kata (3655 words)
master_rows = list(master_ws.iter_rows(values_only=True))[1:]
print(f"Processing {len(master_rows)} master words...")

words_list = []

for row in master_rows:
    if not row or not row[0] or not row[5]:
        continue

    try:
        no = int(row[0])
    except (ValueError, TypeError):
        continue

    tanggal = str(row[1]) if len(row) > 1 and row[1] else ''
    minggu = int(row[2]) if len(row) > 2 and row[2] else 1
    sumber = str(row[3]) if len(row) > 3 and row[3] else 'Vocabulary'
    prioritas = str(row[4]) if len(row) > 4 and row[4] else ''
    word_str = str(row[5]).strip()
    word_lower = word_str.lower()
    
    ipa = str(row[6]) if len(row) > 6 and row[6] else ''
    ipa_perkiraan = str(row[7]) if len(row) > 7 and row[7] else ''
    meaning_id = str(row[8]) if len(row) > 8 and row[8] else ''
    example = str(row[9]) if len(row) > 9 and row[9] else ''
    synonyms_raw = str(row[10]) if len(row) > 10 and row[10] else ''
    antonyms_raw = str(row[11]) if len(row) > 11 and row[11] else ''

    # Get from Word Family map
    wf = wf_map.get(word_lower, {})
    pos = wf.get('pos') or ('verb' if 'verb' in prioritas.lower() else 'noun')
    noun_f = wf.get('noun_f', '-')
    verb_f = wf.get('verb_f', '-')
    adj_f = wf.get('adj_f', '-')
    adv_f = wf.get('adv_f', '-')
    colloc_raw = wf.get('colloc', '-')

    # Get from Afiks map
    afiks = afiks_map.get(word_lower, [])

    # Cross-fill missing word families using Afiks map
    derived_nouns = [item['derived'] for item in afiks if item['pos'] in ['n', 'noun']]
    derived_verbs = [item['derived'] for item in afiks if item['pos'] in ['v', 'verb']]
    derived_adjs = [item['derived'] for item in afiks if item['pos'] in ['adj', 'adjective']]
    derived_advs = [item['derived'] for item in afiks if item['pos'] in ['adv', 'adverb']]

    if noun_f == '-' and derived_nouns:
        noun_f = ', '.join([f"{n} (n)" for n in derived_nouns])
    if verb_f == '-' and derived_verbs:
        verb_f = ', '.join([f"{v} (v)" for v in derived_verbs])
    if adj_f == '-' and derived_adjs:
        adj_f = ', '.join([f"{a} (adj)" for a in derived_adjs])
    if adv_f == '-' and derived_advs:
        adv_f = ', '.join([f"{a} (adv)" for a in derived_advs])

    # Construct Prefix and Suffix info
    prefix_items = [f"{item['used']} ({item['derived']})" for item in afiks if item['type'] == 'Prefix']
    suffix_items = [f"{item['used']} ({item['derived']})" for item in afiks if item['type'] == 'Suffix']

    prefix_info = ', '.join(prefix_items) if prefix_items else ''
    suffix_info = ', '.join(suffix_items) if suffix_items else ''

    # Clean Collocations
    if colloc_raw and colloc_raw != '-':
        collocations = [c.strip() for c in colloc_raw.split(',') if c.strip()]
    else:
        collocations = []

    # Clean Synonyms & Antonyms
    synonyms = [s.strip() for s in synonyms_raw.replace(';', ',').split(',') if s.strip() and s.strip() != '-']
    antonyms = [a.strip() for a in antonyms_raw.replace(';', ',').split(',') if a.strip() and a.strip() != '-']

    word_entry = {
        'id': no,
        'no': no,
        'word': word_str,
        'ipa': ipa,
        'ipa_perkiraan': ipa_perkiraan,
        'meaning_id': meaning_id,
        'pos': pos,
        'source': sumber,
        'priority': prioritas,
        'v1': word_str if 'verb' in pos.lower() else '-',
        'v2': '-',
        'v3': '-',
        'v_ing': '-',
        'verb_type': 'transitive',
        'noun_family': noun_f,
        'verb_family': verb_f,
        'adj_family': adj_f,
        'adv_family': adv_f,
        'collocations': collocations,
        'synonyms': synonyms,
        'antonyms': antonyms,
        'example': example,
        'prefix_info': prefix_info,
        'suffix_info': suffix_info
    }

    words_list.append(word_entry)

print(f"Total processed words: {len(words_list)}")

# Inspect happy specifically
happy_entry = next((w for w in words_list if w['word'] == 'happy'), None)
print("Processed 'happy' entry:")
print(json.dumps(happy_entry, indent=2))

# Write to wordsMaster.ts
ts_content = f"""import {{ Word }} from '../types';

export const MASTER_WORDS: Word[] = {json.dumps(words_list, indent=2)};

export function getWordById(id: number): Word | undefined {{
  return MASTER_WORDS.find((w) => w.id === id);
}}

export function getWordsRange(startNo: number, endNo: number): Word[] {{
  return MASTER_WORDS.filter((w) => w.no >= startNo && w.no <= endNo);
}}
"""

with open(target_ts_path, 'w', encoding='utf-8') as f:
    f.write(ts_content)

print(f"Successfully updated {target_ts_path}!")
