import openpyxl
import json
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

excel_path = os.path.join(os.path.dirname(__file__), '..', 'source', 'Skema_Belajar_5Bulan_Lengkap.xlsx')
output_json_path = os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'words3655.json')

print(f"Loading workbook from {excel_path}...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

ws_master = wb['Master Kata (Bernomor)']
ws_family = wb['Word Family & Kolokasi']

rows_master = list(ws_master.iter_rows(values_only=True))
rows_family = list(ws_family.iter_rows(values_only=True))

family_dict = {}
for r in rows_family[1:]:
    if not r or r[0] is None:
        continue
    no = int(r[0])
    family_dict[no] = {
        'word': r[1],
        'pos': r[2],
        'noun_family': r[3] if r[3] is not None else '-',
        'verb_family': r[4] if r[4] is not None else '-',
        'adj_family': r[5] if r[5] is not None else '-',
        'adv_family': r[6] if r[6] is not None else '-',
        'collocations': r[7] if r[7] is not None else '-'
    }

words = []
for r in rows_master[1:]:
    if not r or r[0] is None:
        continue
    no = int(r[0])
    tgl = str(r[1]) if r[1] is not None else ''
    minggu = int(r[2]) if r[2] is not None else 1
    sumber = str(r[3]).strip() if r[3] is not None else 'Vocabulary'
    prioritas = str(r[4]).strip() if r[4] is not None else 'Tinggi'
    kata = str(r[5]).strip() if r[5] is not None else ''
    ipa = str(r[6]).strip() if r[6] is not None else ''
    ipa_perkiraan = str(r[7]).strip() if r[7] is not None else ''
    arti = str(r[8]).strip() if r[8] is not None else ''
    contoh = str(r[9]).strip() if r[9] is not None else ''
    sinonim_str = str(r[10]).strip() if r[10] is not None else ''
    antonim_str = str(r[11]).strip() if r[11] is not None else ''

    fam = family_dict.get(no, {})

    # Determine POS
    pos_raw = fam.get('pos')
    pos = 'Noun'
    if pos_raw:
        pos_str = str(pos_raw).strip()
        if pos_str.lower() == 'noun':
            pos = 'Noun'
        elif pos_str.lower() == 'verb':
            pos = 'Verb'
        elif pos_str.lower() in ['adj', 'adjective']:
            pos = 'Adjective'
        elif pos_str.lower() in ['adv', 'adverb']:
            pos = 'Adverb'
        elif pos_str.lower() in ['phrasal verb', 'phrasal']:
            pos = 'Phrasal Verb'
        elif pos_str.lower() == 'idiom':
            pos = 'Idiom'
        else:
            pos = pos_str.capitalize()

    synonyms = [s.strip() for s in sinonim_str.split(',') if s.strip() and s.strip() != '-']
    antonyms = [a.strip() for a in antonim_str.split(',') if a.strip() and a.strip() != '-']
    
    colloc_raw = str(fam.get('collocations', '')).strip()
    collocations = [c.strip() for c in colloc_raw.split(',') if c.strip() and c.strip() != '-']

    # Verb forms calculation/extraction if Verb
    v1 = kata if pos == 'Verb' else None
    v2 = f"{kata}d" if pos == 'Verb' and not kata.endswith('ed') else (f"{kata}" if pos == 'Verb' else None)
    v3 = v2
    v_ing = f"{kata[:-1]}ing" if pos == 'Verb' and kata.endswith('e') else (f"{kata}ing" if pos == 'Verb' else None)

    # Word family values
    noun_fam = str(fam.get('noun_family', '-')).strip()
    verb_fam = str(fam.get('verb_family', '-')).strip()
    adj_fam = str(fam.get('adj_family', '-')).strip()
    adv_fam = str(fam.get('adv_family', '-')).strip()

    word_entry = {
        'id': no,
        'no': no,
        'word': kata,
        'pos': pos,
        'ipa': ipa if ipa and ipa != '-' else '',
        'ipa_perkiraan': ipa_perkiraan if ipa_perkiraan and ipa_perkiraan != '-' else '',
        'meaning_id': arti if arti and arti != '-' else '',
        'example_sentence': contoh if contoh and contoh != '-' else f"The word '{kata}' is widely used in standard English.",
        'source': sumber,
        'priority': prioritas,
        'synonyms': synonyms,
        'antonyms': antonyms,
        'collocations': collocations,
        'noun_family': noun_fam if noun_fam != '' else '-',
        'verb_family': verb_fam if verb_fam != '' else '-',
        'adj_family': adj_fam if adj_fam != '' else '-',
        'adv_family': adv_fam if adv_fam != '' else '-',
        'week': minggu
    }

    if pos == 'Verb':
        word_entry['v1'] = v1
        word_entry['v2'] = v2
        word_entry['v3'] = v3
        word_entry['v_ing'] = v_ing
        word_entry['verb_type'] = 'transitive'

    words.append(word_entry)

os.makedirs(os.path.dirname(output_json_path), exist_ok=True)
with open(output_json_path, 'w', encoding='utf-8') as f:
    json.dump(words, f, ensure_ascii=False, indent=2)

print(f"Successfully extracted {len(words)} words to {output_json_path}")
